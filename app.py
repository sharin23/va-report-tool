from flask import Flask, request, send_file, render_template_string
import pandas as pd
import xml.etree.ElementTree as ET
from io import BytesIO
from openpyxl.styles import PatternFill
import os
import zipfile

app = Flask(__name__)

# ===== HTML template with dashboard-style tabs =====
HOME_PAGE = """
<!doctype html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>VA Report Tool</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.2/css/all.min.css" rel="stylesheet">
    <style>
        body { background-color: #eef2f7; padding-top: 50px; }
        .container { max-width: 900px; }
        h1 { text-align: center; color: #1f2a38; margin-bottom: 40px; font-weight: 700; }
        .card { border-radius: 15px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); margin-top: 20px; }
        .card-title { font-weight: 600; color: #0d6efd; }
        .card-text { font-size: 0.95rem; color: #495057; }
        .btn-primary, .btn-success { width: 100%; font-weight: 500; }
        label { font-weight: 500; }
        footer { font-size: 0.85rem; color: #6c757d; }
        .mode-select { transition: all 0.3s ease-in-out; }
    </style>
    <script>
        function toggleModeSelect() {
            var files = document.getElementById('nessus_files').files;
            var modeDiv = document.getElementById('mode_div');
            modeDiv.style.display = files.length > 1 ? 'block' : 'none';
        }
    </script>
</head>
<body>
<div class="container">
    <h1><i class="fas fa-shield-alt"></i> Vulnerability Report Manager </h1>

    <!-- Tabs -->
    <ul class="nav nav-tabs" id="vaTab" role="tablist">
      <li class="nav-item" role="presentation">
        <button class="nav-link active" id="convert-tab" data-bs-toggle="tab" data-bs-target="#convert" type="button" role="tab" aria-controls="convert" aria-selected="true">
            <i class="fas fa-file-excel"></i> Convert Nessus
        </button>
      </li>
      <li class="nav-item" role="presentation">
        <button class="nav-link" id="compare-tab" data-bs-toggle="tab" data-bs-target="#compare" type="button" role="tab" aria-controls="compare" aria-selected="false">
            <i class="fas fa-exchange-alt"></i> Compare Reports
        </button>
      </li>
    </ul>

    <div class="tab-content">
        <!-- Convert Nessus Tab -->
        <div class="tab-pane fade show active" id="convert" role="tabpanel" aria-labelledby="convert-tab">
            <div class="card shadow-sm">
                <div class="card-body">
                    <h4 class="card-title">Convert .nessus Files</h4>
                    <p class="card-text">
                        Upload one or more .nessus files. For multiple files, choose whether to merge into a single Excel or convert each separately.
                    </p>
                    <form method="post" enctype="multipart/form-data" action="/convert_nessus">
                        <div class="mb-3">
                            <input class="form-control" type="file" id="nessus_files" name="nessus_files" multiple required onchange="toggleModeSelect()">
                        </div>
                        <div class="mb-3 mode-select" id="mode_div" style="display:none;">
                            <label>Conversion Mode (for multiple files):</label>
                            <select class="form-control" name="mode">
                                <option value="merge">Merge Multiple Nessus Files → Single Excel</option>
                                <option value="multi">Multiple Nessus Files → Separate Excel Files (ZIP)</option>
                            </select>
                        </div>
                        <button type="submit" class="btn btn-primary"><i class="fas fa-play"></i> Generate Report</button>
                    </form>
                </div>
            </div>
        </div>

        <!-- Compare Reports Tab -->
        <div class="tab-pane fade" id="compare" role="tabpanel" aria-labelledby="compare-tab">
            <div class="card shadow-sm">
                <div class="card-body">
                    <h4 class="card-title">Compare Reports</h4>
                    <p class="card-text">
                        Upload two Excel reports to generate a combined vulnerability report, highlighting Recurring, Fixed, and New issues.
                    </p>
                    <form method="post" enctype="multipart/form-data" action="/compare_reports">
                        <div class="mb-3">
                            <label>Previous Report:</label>
                            <input class="form-control" type="file" name="q1_file" required>
                        </div>
                        <div class="mb-3">
                            <label>Current Report:</label>
                            <input class="form-control" type="file" name="q2_file" required>
                        </div>
                        <button type="submit" class="btn btn-success"><i class="fas fa-chart-line"></i> Compare Reports</button>
                    </form>
                </div>
            </div>
        </div>
    </div>

    <footer class="text-center mt-4 mb-4">
        Developed by <strong>HPSL</strong> &nbsp; | &nbsp; <i class="fas fa-copyright"></i> 2025
    </footer>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

# ===== Helper: Parse Nessus File =====
def parse_nessus_file(nessus_file):
    severity_map = {"1": "Low", "2": "Medium", "3": "High", "4": "Critical"}
    tree = ET.parse(nessus_file)
    root = tree.getroot()
    findings = []

    for report_host in root.findall(".//ReportHost"):
        host = report_host.attrib.get("name", "Unknown Host")

        for report_item in report_host.findall("ReportItem"):
            severity = report_item.attrib.get("severity")
            if severity == "0":
                continue

            plugin_name = report_item.attrib.get("pluginName", "Unknown")
            severity_word = severity_map.get(severity, "Unknown")
            description = report_item.findtext("description", "")
            solution = report_item.findtext("solution", "")
            synopsis = report_item.findtext("synopsis", "")
            port = report_item.attrib.get("port", "")

            findings.append({
                "Host": host,
                "Severity": severity_word,
                "Port": port,
                "Vulnerability Name": plugin_name,
                "Description": description,
                "Synopsis": synopsis,
                "Solution": solution
            })

    return pd.DataFrame(findings)

# ===== Severity color map =====
COLOR_MAP = {
    "Critical": "800080",
    "High": "FF0000",
    "Medium": "FFA500",
    "Low": "00FF00"
}

# ===== Convert Nessus =====
@app.route("/convert_nessus", methods=["POST"])
def convert_nessus():
    files = request.files.getlist("nessus_files")
    mode = request.form.get("mode")
    if not files:
        return "No files uploaded!"

    def process_df(df):
        df.drop_duplicates(subset=["Host", "Vulnerability Name", "Port"], inplace=True)
        if "Sr No" in df.columns:
            df.drop(columns=["Sr No"], inplace=True)
        severity_order = ["Critical", "High", "Medium", "Low"]
        df["Severity"] = pd.Categorical(df["Severity"], categories=severity_order, ordered=True)
        df.sort_values(by="Severity", inplace=True)
        df.insert(0, "Sr No", range(1, len(df)+1))
        return df

    # SINGLE FILE
    if len(files) == 1:
        df = process_df(parse_nessus_file(files[0]))
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Report")
            ws = writer.sheets["Report"]
            severity_idx = df.columns.get_loc("Severity")+1
            for row in range(2, len(df)+2):
                cell = ws.cell(row=row, column=severity_idx)
                color = COLOR_MAP.get(cell.value, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name=f"{os.path.splitext(files[0].filename)[0]}.xlsx")

    # MULTIPLE FILES
    if not mode:
        return "Please select a conversion mode for multiple files."

    if mode == "merge":
        all_dfs = [process_df(parse_nessus_file(f)) for f in files]
        df = pd.concat(all_dfs, ignore_index=True)
        df = process_df(df)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Merged_Report")
            ws = writer.sheets["Merged_Report"]
            severity_idx = df.columns.get_loc("Severity")+1
            for row in range(2, len(df)+2):
                cell = ws.cell(row=row, column=severity_idx)
                color = COLOR_MAP.get(cell.value, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="Merged_Nessus_Report.xlsx")

    elif mode == "multi":
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for f in files:
                df = process_df(parse_nessus_file(f))
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Report")
                    ws = writer.sheets["Report"]
                    severity_idx = df.columns.get_loc("Severity")+1
                    for row in range(2, len(df)+2):
                        cell = ws.cell(row=row, column=severity_idx)
                        color = COLOR_MAP.get(cell.value, "FFFFFF")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                excel_buffer.seek(0)
                zip_file.writestr(f"{os.path.splitext(f.filename)[0]}.xlsx", excel_buffer.read())
        zip_buffer.seek(0)
        return send_file(zip_buffer, as_attachment=True, download_name="Multiple_Nessus_Reports.zip")

    else:
        return "Invalid conversion mode!"

# ===== Compare Reports (Host + Vulnerability + Port) =====
@app.route("/compare_reports", methods=["POST"])
def compare_reports():
    q1 = request.files.get("q1_file")
    q2 = request.files.get("q2_file")
    if not q1 or not q2:
        return "Both files are required!"

    df_q1 = pd.read_excel(q1)
    df_q2 = pd.read_excel(q2)
    df_q1.columns = df_q1.columns.str.strip()
    df_q2.columns = df_q2.columns.str.strip()

    vuln_col = next((c for c in ["Vulnerability Name","Plugin Name","Vulnerability"]
                     if c in df_q1.columns and c in df_q2.columns), None)
    if not vuln_col:
        return "No common vulnerability column found!"

    # Create unique key based on Host + Vulnerability + Port
    df_q1["UniqueKey"] = df_q1["Host"].astype(str) + "_" + df_q1[vuln_col].astype(str) + "_" + df_q1["Port"].astype(str)
    df_q2["UniqueKey"] = df_q2["Host"].astype(str) + "_" + df_q2[vuln_col].astype(str) + "_" + df_q2["Port"].astype(str)

    df_q1["Status"] = df_q1["UniqueKey"].apply(lambda x: "Recurring" if x in df_q2["UniqueKey"].values else "Fixed")
    df_q2["Status"] = df_q2["UniqueKey"].apply(lambda x: "Recurring" if x in df_q1["UniqueKey"].values else "New")

    combined_df = pd.concat([df_q1, df_q2[df_q2["Status"]=="New"]], ignore_index=True)
    combined_df.drop(columns=["UniqueKey"], inplace=True)

    if "Severity" in combined_df.columns:
        valid = ["Critical","High","Medium","Low"]
        combined_df = combined_df[combined_df["Severity"].isin(valid)]
        combined_df["Severity"] = pd.Categorical(combined_df["Severity"], categories=valid, ordered=True)
        combined_df.sort_values("Severity", inplace=True)

    # Remove existing Sr No if present, then insert
    if "Sr No" in combined_df.columns:
        combined_df.drop(columns=["Sr No"], inplace=True)
    combined_df.insert(0, "Sr No", range(1, len(combined_df)+1))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        combined_df.to_excel(writer, index=False, sheet_name="Comparison")
        ws = writer.sheets["Comparison"]
        if "Severity" in combined_df.columns:
            severity_idx = combined_df.columns.get_loc("Severity")+1
            for row in range(2, len(combined_df)+2):
                cell = ws.cell(row=row, column=severity_idx)
                color = COLOR_MAP.get(cell.value, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    output.seek(0)
    return send_file(output, as_attachment=True, download_name="Combined_Report.xlsx")

# ===== Homepage =====
@app.route("/")
def home():
    return render_template_string(HOME_PAGE)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)

